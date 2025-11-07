#!/usr/bin/env python3
"""
Script to transform MATRIZ.xlsx into a proper import file with database ID references.

This script:
1. Reads the PROGRAMACION sheet from MATRIZ.xlsx
2. Maps Sub Area, Cargo, and Tema names to their database IDs
3. Looks up Rol from the DIM sheet
4. Creates a new Excel file ready for import: matriz_importable.xlsx

The output format is:
Column A: ID Cargo (e.g., 117)
Column B: ID Sub Área (e.g., 003)
Column C: ID Tema (number)
Column D: Frecuencia en meses
Column E: Nombre del Rol Capacitador
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Database mappings extracted from osm_postgres.sql

# Sub Areas mapping (name -> id_area)
SUB_AREAS_DB = {
    'LOGISTICA Y COSECHA': '001',
    'PRODUCCION AGRICOLA': '002',
    'MANTENIMIENTO INDUSTRIAL': '003',
    'FRUTA PROVEEDORES': '004',
    'ADMINISTRACION': '005',
    'ALMACEN': '006',
    'AGRONOMICA': '007',
    'SANIDAD': '008',
    'POLINIZACION': '009',
    'COMPRAS': '010',
    'CLIPA - SISTEMAS': '011',
    'CONTABILIDAD': '012',
    'VENTAS': '013',
    'BASCULA': '014',
    'LABORATORIO': '015',
    'PLANTA EXTRACTORA': '016',
    'COMPOST': '017',
}

# Cargos mapping (name -> id_cargo)
CARGOS_DB = {
    'ANALISTA AGRONOMICO': '141',
    'DIBUJANTE TECNICO': '135',
    'LIDER ANALISTA': '122',
    'TECNOLOGO AGRICOLA': '137',
    'TESORERO': '011',
    'ANALISTA DE INFORMACION': '116',
    'ASISTENTE ADMNISTRATIVO': '027',
    'ASISTENTE CONTABLE': '114',
    'ASISTENTE DE ALMACEN Y COMPRAS': '014',
    'ASISTENTE DE MANTENIMIENTO': '100',
    'ASISTENTE DE SISTEMAS': '017',
    'ASISTENTE PRODUCCION INDUSTRIAL': '125',
    'PLANEADOR MANTENIMIENTO': '119',
    'AUXILIAR ADMINISTRATIVO': '085',
    'AUXILIAR CONTABLE': '013',
    'AUXILIAR DE LABORATORIO': '052',
    'AUXILIAR DE SUPERVISION AGRICOLA': '030',
    'CONDUCTOR': '024',
    'CONDUCTOR -MENSAJERO': '018',
    'COORDINADOR ADMINISTRATIVO': '129',
    'COORDINADOR CONTABLE': '081',
    'COORDINADOR DE PROYECTOS': '105',
    'INGENIERA AGRICOLA': '128',
    'COORDINADOR DE LOGISTICA': '091',
    'COORDINADOR PRODUCCION INDUSTRIAL': '076',
    'COORDINADOR AGRONOMO': '136',
    'DIRECTOR DE MANTENIMIENTO INDUSTRIAL': '117',
    'DIRECTOR DE PLANTA': '043',
    'DIRECTOR LOGISTICA, COSECHA Y MANTENIMIENTO': '068',
    'DIRECTORA ADMINISTRATIVA Y FINANCIERA': '127',
    'JEFE DE COMPRAS': '007',
    'JEFE DE VENTAS': '126',
    'AUXILIAR DE APOYO': '140',
    'AUXILIAR DE SERVICIOS GENERALES I': '020',
    'AUXILIAR DE MANTENIMIENTO Y SOLDADURA': '054',
    'ELECTROMECANICO': '046',
    'OPERARIO ENERGIA': '118',
    'SOLDADOR MECANICO': '049',
    'OPERARIO DE MAQUINARIA AGRICOLA': '031',
    'OPERARIO DE PLANTA I': '055',
    'OPERARIO DE PLANTA II': '058',
    'OFICIOS VARIOS II': '033',
    'OFICIOS VARIOS III': '029',
    'OFICIOS VARIOS IV': '095',
    'OPERARIO DE LOGISTICA Y COSECHA': '104',
    'POLINIZADOR': '073',
    'SUPERVISOR DE LABORATORIO': '130',
    'SUPERVISOR DE LOGISTICA I': '089',
    'SUPERVISOR DE PROCESO': '047',
    'SUPERVISOR DE PRODUCCION': '072',
    'SUPERVISOR DE MANTENIMIENTO INDUSTRIAL': '112',
    'SUPERVISOR DE TOPOGRAFIA': '123',
    'DIRECTOR DE PROYECTOS': '004',
    'COORDINADOR DE GESTION HUMANA': '008',
    'ASISTENTE DE SEGURIDAD Y SALUD EN EL TRABAJO': '016',
    'AUXILIAR SERVICIOS GENERALES II': '078',
    'DIRECTORA CONTABLE': '005',
    'GERENTE ADMINISTRATIVO': '003',
    'COORDINADOR DE SISTEMAS': '006',
    'GERENTE GENERAL': '001',
    'ANALISTA CONTABLE  1': '009',
    'MENSAJERO': '019',
    'ASISTENTE ADMINISTRATIVO II': '077',
    'COORDINADOR AMBIENTAL Y SO': '010',
    'ASISTENTE ADMINISTRATIVA I': '015',
    'ESTUDIANTE SENA ETAPA PRODUCTIVA': '021',
    'ESTUDIANTE SENA ETAPA ELECTIVA': '040',
    'COORDINADOR DE SEGURIDAD Y SALUD EN EL TRABAJO': '093',
    'AUXILIAR DE SISTEMAS': '084',
    'AUXILIAR ADMINISTRATIVAIII': '106',
    'COORDINADOR DE ANALITICA Y SISTEMAS': '143',
}

# Temas mapping (name -> id)
TEMAS_DB = {
    'Trabajo en equipo ': 1,
    'Trabajo en equipo': 1,
    'Capacitación a supervisores de proceso': 2,
    'Principios y criterios RSPO ': 3,
    'Principios y criterios RSPO': 3,
    'Inducción Talento humano y gestión administrativa': 4,
    'Procedimiento de despacho': 5,
    'Procedimiento de producto No conforme': 6,
    'Calidad de aceite (Índice en yodo)': 7,
    'Perdidas de almendra': 8,
    'Manual de cadena de suministro y socialización de responsabilidades. ': 9,
    'Manual de cadena de suministro y socialización de responsabilidades': 9,
    'Llenado de vagones': 10,
    'Mesa de volteo': 11,
    'Esterilización ': 12,
    'Esterilización': 12,
    'Desfrutado, digestión, prensado y desfibrado.': 13,
    'Desfrutado, digestión, prensado y desfibrado': 13,
    'Clarificación ': 14,
    'Clarificación': 14,
    'Almacenamiento y control de tanques de almacenamiento': 15,
    'Palmisteria': 16,
    'Caldera': 17,
    'Proceso de extracción de aceite': 18,
    'Procedimiento de recepción de RFF': 19,
    'Instructivo para operar bascula': 20,
    'Descargue y calificación de fruta': 21,
    'Procedimiento de orden y aseo': 22,
    'BPM e inocuidad alimentaria': 23,
    'Manejo de plaguicidas ': 24,
    'Manejo de plaguicidas': 24,
    'Socialización de instructivos - Monitoreos ': 25,
    'Socialización de instructivos - Monitoreos': 25,
    'Socialización de instructivos - Polinización': 26,
    'Socialización de instructivos - logística de producción': 27,
    'Socialización de ISCC,RSPO, ORGANICO': 28,
    'Requerimientos Cadena de suministro RSPO': 29,
    'Socialización de procedimiento PQRS': 30,
    'Política de sostenibilidad': 31,
    'MIPE': 32,
    'Socialización de instructivos - Fertilización': 33,
    'Manejo de Insumos orgánicos': 34,
    'Manejo defensivo': 35,
    'Manejo de cybertracker para el envió de información.': 36,
    'Manejo de cybertracker para el envió de información': 36,
    'Supervisión labores de corte de RFF.': 37,
    'Supervisión labores de corte de RFF': 37,
    'Supervisión de labores de mantenimiento en campo en palma de aceite.': 38,
    'Supervisión de labores de mantenimiento en campo en palma de aceite': 38,
    'Supervisión de labores de fertilización en palma de aceite.': 39,
    'Supervisión de labores de fertilización en palma de aceite': 39,
    'Logística para cosecha de lotes orgánicos': 40,
    'Uso eficiente de maquinaria agrícola y aplicaciones de agricultura de precisión': 41,
    'Riesgo biológico': 42,
    'Uso adecuado EPP/Sensibilización de limpieza y mantenimiento de epp': 43,
    'Reporte de accidentes laborales, incidentes, y EL': 44,
    'Identificación de riesgos y peligros': 45,
    'Reporte de actos y condiciones inseguras': 46,
    'Autocuidado ': 47,
    'Autocuidado': 47,
    'Identificación de peligros': 48,
    'Riesgo Psicosocial': 49,
    'Manejo racional de plaguicidas': 50,
    'Riesgo químico. SGA': 51,
    'Riesgo ergonómico': 52,
    'Inducción y reinducción de seguridad y salud en el trabajo ': 53,
    'Inducción y reinducción de seguridad y salud en el trabajo': 53,
    'Brigadistas de emergencia': 54,
    'Brigadas de emergencia': 54,
    'Capacitación al COPASST ': 55,
    'Capacitación al COPASST': 55,
    'Curso de 50 horas/ 20 horas': 56,
    'Comunicación asertiva': 57,
    'Excel': 58,
    'Análisis e datos': 59,
    'Calibración y Mantenimiento de Equipos': 60,
    'Manejo del suelo ': 61,
    'Manejo del suelo': 61,
    'Control de calidad': 62,
    'Diseño Web': 63,
    'Electricidad industrial ': 64,
    'Electricidad industrial': 64,
    'Generación de energía': 65,
    'Herramientas de Google': 66,
    'Herramientas ofimáticas': 67,
    'Manejo de aplicativo ATLAS': 68,
    'Manejo de maquinaria y equipo': 69,
    'Manejo integrado de plagas y enfermedades': 70,
    'Mantenimiento de cámaras': 71,
    'Mantenimientos de primer nivel ': 72,
    'Mantenimientos de primer nivel': 72,
    'Nomina y seguridad social ': 73,
    'Nomina y seguridad social': 73,
    'Planeación de labores agrícolas': 74,
    'Power Bi': 75,
    'Rotación de inventario': 76,
    'SAP': 77,
    'Bienestar laboral': 78,
    'Riesgo vial': 79,
    'Trabajo seguro en alturas': 80,
    'Trabajo en espacios confinados ': 81,
    'Trabajo en espacios confinados': 81,
    'Coordinador trabajo seguro en alturas': 80,  # Likely the same as #80
    'Actualizacion tributaria': 82,  # Note: ID 82 needs to be added to cap_tema table
    'Manejo racional de plaguicidas (Agroquimicos)': 50,  # Maps to existing "Manejo racional de plaguicidas"
}


def normalize_text(text):
    """Normalize text for comparison."""
    if not text:
        return ''
    return str(text).strip().upper()


def load_dim_mapping(wb):
    """Load the DIM sheet to create a tema -> rol mapping."""
    tema_to_rol = {}
    try:
        ws = wb['DIM ']  # Note the space
        for row in ws.iter_rows(min_row=3, values_only=True):
            tema_name = row[0]
            rol_name = row[2]
            if tema_name and rol_name:
                tema_to_rol[str(tema_name).strip()] = str(rol_name).strip()
    except Exception as e:
        print(f"Warning: Could not load DIM sheet: {e}")
    return tema_to_rol


def transform_matriz():
    """Main transformation function."""
    print("Loading MATRIZ.xlsx...")
    wb = openpyxl.load_workbook('MATRIZ.xlsx')
    
    # Load tema -> rol mapping from DIM sheet
    print("Loading DIM sheet mapping...")
    tema_to_rol = load_dim_mapping(wb)
    
    # Read PROGRAMACION sheet
    ws_prog = wb['PROGRAMACION']
    print(f"Found {ws_prog.max_row} rows in PROGRAMACION sheet")
    
    # Create new workbook for output
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "programacion"
    
    # Write header
    headers = ['Cargo ID', 'Sub Área ID', 'Tema ID', 'Frecuencia', 'Rol Capacitador']
    for col, header in enumerate(headers, 1):
        cell = ws_out.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Track statistics
    stats = {
        'total': 0,
        'success': 0,
        'missing_cargo': 0,
        'missing_sub_area': 0,
        'missing_tema': 0,
        'missing_rol': 0
    }
    
    errors = []
    output_row = 2
    
    # Process each row from the PROGRAMACION sheet
    for row_num, row in enumerate(ws_prog.iter_rows(min_row=3, values_only=True), 3):
        stats['total'] += 1
        
        # Extract data
        sub_area_name = str(row[0]).strip() if row[0] else ''
        cargo_name = str(row[1]).strip() if row[1] else ''
        tema_name = str(row[2]).strip() if row[2] else ''
        frecuencia = row[3] if row[3] and not str(row[3]).startswith('=') else 12
        
        # Skip empty rows
        if not sub_area_name or not cargo_name or not tema_name:
            continue
        
        # Map to database IDs
        sub_area_id = SUB_AREAS_DB.get(normalize_text(sub_area_name))
        cargo_id = CARGOS_DB.get(normalize_text(cargo_name))
        tema_id = TEMAS_DB.get(tema_name)
        
        # Get rol from DIM mapping
        rol_name = tema_to_rol.get(tema_name, '')
        
        # Track missing mappings
        if not cargo_id:
            stats['missing_cargo'] += 1
            errors.append(f"Row {row_num}: Cargo not found: '{cargo_name}'")
            continue
        
        if not sub_area_id:
            stats['missing_sub_area'] += 1
            errors.append(f"Row {row_num}: Sub Area not found: '{sub_area_name}'")
            continue
        
        if not tema_id:
            stats['missing_tema'] += 1
            errors.append(f"Row {row_num}: Tema not found: '{tema_name}'")
            continue
        
        if not rol_name:
            stats['missing_rol'] += 1
            errors.append(f"Row {row_num}: Rol not found for tema: '{tema_name}'")
            # Don't skip, we'll still output the row but with empty rol
        
        # Write to output
        ws_out.cell(row=output_row, column=1, value=cargo_id)
        ws_out.cell(row=output_row, column=2, value=sub_area_id)
        ws_out.cell(row=output_row, column=3, value=tema_id)
        ws_out.cell(row=output_row, column=4, value=int(frecuencia))
        ws_out.cell(row=output_row, column=5, value=rol_name)
        
        output_row += 1
        stats['success'] += 1
    
    # Auto-adjust column widths
    for col in range(1, 6):
        ws_out.column_dimensions[get_column_letter(col)].width = 20
    
    # Save output file
    output_file = 'matriz_importable.xlsx'
    wb_out.save(output_file)
    
    # Print statistics
    print("\n" + "="*60)
    print("TRANSFORMATION COMPLETE")
    print("="*60)
    print(f"Total rows processed: {stats['total']}")
    print(f"Successfully transformed: {stats['success']}")
    print(f"Missing cargo mappings: {stats['missing_cargo']}")
    print(f"Missing sub_area mappings: {stats['missing_sub_area']}")
    print(f"Missing tema mappings: {stats['missing_tema']}")
    print(f"Missing rol mappings: {stats['missing_rol']}")
    print(f"\nOutput file: {output_file}")
    
    if errors:
        print("\n" + "="*60)
        print("ERRORS/WARNINGS (first 20):")
        print("="*60)
        for error in errors[:20]:
            print(error)
        if len(errors) > 20:
            print(f"\n... and {len(errors) - 20} more errors")
    
    print("\n" + "="*60)


if __name__ == '__main__':
    transform_matriz()
