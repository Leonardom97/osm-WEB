#!/usr/bin/env python3
"""
Test script to validate the import file format and data integrity.
This script simulates the validation that happens in the web interface.
"""

import openpyxl

# Load the database mappings for validation
# These should match the actual database content
VALID_CARGO_IDS = ['001', '003', '004', '005', '006', '007', '008', '009', '010', '011', 
                   '013', '014', '015', '016', '017', '018', '019', '020', '021', '024',
                   '027', '029', '030', '031', '033', '040', '043', '046', '047', '049',
                   '052', '054', '055', '058', '068', '072', '073', '076', '077', '078',
                   '081', '084', '085', '089', '091', '093', '095', '100', '104', '105',
                   '106', '112', '114', '116', '117', '118', '119', '122', '123', '125',
                   '126', '127', '128', '129', '130', '135', '136', '137', '140', '141', '143']

VALID_SUB_AREA_IDS = ['000', '001', '002', '003', '004', '005', '006', '007', '008', '009',
                      '010', '011', '012', '013', '014', '015', '016', '017', '999']

VALID_TEMA_IDS = list(range(1, 82))  # IDs 1-81 based on database

VALID_ROLES = ['Capacitador SIE', 'Capacitador GH', 'Capacitador TI', 'Capacitador MT',
               'Capacitador IND', 'Capacitador ADM', 'Capacitador AGR', 
               'Administrador', 'Capacitador', 'Aux_Capacitador']


def validate_import_file(filename):
    """Validate an Excel file for import."""
    print(f"Validating {filename}...")
    print("=" * 60)
    
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        
        print(f"Sheet name: {ws.title}")
        print(f"Total rows: {ws.max_row}")
        print()
        
        # Check header
        expected_headers = ['Cargo ID', 'Sub Área ID', 'Tema ID', 'Frecuencia', 'Rol Capacitador']
        actual_headers = [cell.value for cell in ws[1]]
        
        if actual_headers[:5] != expected_headers:
            print("❌ ERROR: Header mismatch!")
            print(f"Expected: {expected_headers}")
            print(f"Got: {actual_headers[:5]}")
            return False
        else:
            print("✓ Headers are correct")
        
        # Validate data rows
        errors = []
        warnings = []
        valid_rows = 0
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row[0]:  # Skip empty rows
                continue
            
            cargo_id = str(row[0]).strip() if row[0] else ''
            sub_area_id = str(row[1]).strip() if row[1] else ''
            tema_id = int(row[2]) if row[2] else 0
            frecuencia = int(row[3]) if row[3] else 12
            rol_nombre = str(row[4]).strip() if row[4] else ''
            
            # Validate each field
            row_valid = True
            
            if not cargo_id:
                errors.append(f"Row {row_num}: Cargo ID is empty")
                row_valid = False
            elif cargo_id not in VALID_CARGO_IDS:
                warnings.append(f"Row {row_num}: Cargo ID '{cargo_id}' may not exist in database")
            
            if not sub_area_id:
                errors.append(f"Row {row_num}: Sub Área ID is empty (required)")
                row_valid = False
            elif sub_area_id not in VALID_SUB_AREA_IDS:
                warnings.append(f"Row {row_num}: Sub Área ID '{sub_area_id}' may not exist in database")
            
            if not tema_id:
                errors.append(f"Row {row_num}: Tema ID is empty")
                row_valid = False
            elif tema_id not in VALID_TEMA_IDS:
                warnings.append(f"Row {row_num}: Tema ID {tema_id} may not exist in database")
            
            if not frecuencia or frecuencia < 1 or frecuencia > 120:
                warnings.append(f"Row {row_num}: Frecuencia {frecuencia} is unusual (should be 1-120 months)")
            
            if not rol_nombre:
                errors.append(f"Row {row_num}: Rol Capacitador is empty")
                row_valid = False
            elif rol_nombre not in VALID_ROLES:
                errors.append(f"Row {row_num}: Rol '{rol_nombre}' not valid. Must be one of: {', '.join(VALID_ROLES)}")
                row_valid = False
            
            if row_valid:
                valid_rows += 1
        
        # Print summary
        print()
        print("=" * 60)
        print("VALIDATION SUMMARY")
        print("=" * 60)
        print(f"Total data rows: {ws.max_row - 1}")
        print(f"Valid rows: {valid_rows}")
        print(f"Errors: {len(errors)}")
        print(f"Warnings: {len(warnings)}")
        
        if errors:
            print()
            print("ERRORS (first 10):")
            for error in errors[:10]:
                print(f"  ❌ {error}")
            if len(errors) > 10:
                print(f"  ... and {len(errors) - 10} more errors")
        
        if warnings:
            print()
            print("WARNINGS (first 10):")
            for warning in warnings[:10]:
                print(f"  ⚠️  {warning}")
            if len(warnings) > 10:
                print(f"  ... and {len(warnings) - 10} more warnings")
        
        if not errors:
            print()
            print("✓ File is valid and ready for import!")
            return True
        else:
            print()
            print("❌ File has errors. Please fix them before importing.")
            return False
            
    except Exception as e:
        print(f"❌ ERROR: Could not validate file: {e}")
        return False


if __name__ == '__main__':
    import sys
    
    files_to_test = [
        'matriz_importable.xlsx',
        'ejemplo_importacion.xlsx'
    ]
    
    for filename in files_to_test:
        print()
        validate_import_file(filename)
        print()
